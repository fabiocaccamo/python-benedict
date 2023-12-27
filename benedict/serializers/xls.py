import fsutil

try:
    from openpyxl import load_workbook
    from xlrd import open_workbook

    xls_installed = True
except ModuleNotFoundError:
    xls_installed = False

from slugify import slugify

from benedict.extras import require_xls
from benedict.serializers.abstract import AbstractSerializer


class XLSSerializer(AbstractSerializer):
    """
    This class describes a xls serializer.
    """

    def __init__(self):
        super().__init__(
            extensions=[
                "xls",
                "xlsx",
                "xlsm",
            ],
        )

    def _get_sheet_index_and_name_from_options(self, **kwargs):
        sheet_index_or_name = kwargs.pop("sheet", 0)
        sheet_index = 0
        sheet_name = ""
        if isinstance(sheet_index_or_name, int):
            sheet_index = sheet_index_or_name
        elif isinstance(sheet_index_or_name, str):
            sheet_name = sheet_index_or_name
        return (sheet_index, sheet_name)

    def _get_sheet_index_by_name(self, sheet_name, sheet_names):
        sheet_names = [slugify(name) for name in sheet_names]
        try:
            sheet_index = sheet_names.index(slugify(sheet_name))
            return sheet_index
        except ValueError as error:
            raise Exception(
                f"Invalid sheet name {sheet_name!r}, sheet not found."
            ) from error

    def _get_sheet_columns_indexes(self, columns_count):
        return list(range(columns_count))

    def _decode_legacy(self, s, **kwargs):
        options = {}
        options["filename"] = s
        options["logfile"] = kwargs.pop("logfile", None)
        options["verbosity"] = kwargs.pop("verbosity", 0) or 0
        options["use_mmap"] = kwargs.pop("use_mmap", False) or False
        options["file_contents"] = kwargs.pop("file_contents", None)

        # load the worksheet
        workbook = open_workbook(**options)

        # get sheet by index or by name
        sheet_index, sheet_name = self._get_sheet_index_and_name_from_options(**kwargs)
        if sheet_name:
            sheet_names = workbook.sheet_names()
            sheet_index = self._get_sheet_index_by_name(sheet_name, sheet_names)
        sheet = workbook.sheet_by_index(sheet_index)
        sheet_columns_range = range(sheet.ncols)

        # get columns
        columns = kwargs.pop("columns", None)
        columns_row = kwargs.pop("columns_row", True)
        columns_standardized = kwargs.pop("columns_standardized", columns is None)
        if not columns:
            if columns_row:
                # if first row is for column names read the names
                # for row in sheet.iter_rows(min_row=1, max_row=1):
                columns = [
                    sheet.cell_value(0, col_index) for col_index in sheet_columns_range
                ]
            else:
                # otherwise use columns indexes as column names
                # for row in sheet.iter_rows(min_row=1, max_row=1):
                columns = self._get_sheet_columns_indexes(sheet_columns_range)

        # standardize column names, eg. "Date Created" -> "date_created"
        if columns_standardized:
            columns = [slugify(column or "", separator="_") for column in columns]

        # build list of dicts, one for each row
        items = []
        items_row_start = 1 if columns_row else 0
        for row_index in range(items_row_start, sheet.nrows):
            row = {}
            for col_index in sheet_columns_range:
                col_key = columns[col_index]
                value = sheet.cell_value(row_index, col_index)
                row[col_key] = value
            items.append(row)

        # print(items)
        return items

    def _decode(self, s, **kwargs):
        options = {}
        options["filename"] = s
        options["read_only"] = True
        options["data_only"] = kwargs.pop("data_only", False)
        options["keep_links"] = kwargs.pop("keep_links", True)
        options["keep_vba"] = kwargs.pop("keep_vba", True)

        # load the worksheet
        workbook = load_workbook(**options)

        # get sheet by index or by name
        sheet_index, sheet_name = self._get_sheet_index_and_name_from_options(**kwargs)
        sheets = list(workbook)
        if sheet_name:
            sheet_names = [sheet.title for sheet in sheets]
            sheet_index = self._get_sheet_index_by_name(sheet_name, sheet_names)
        sheet = sheets[sheet_index]
        sheet_columns_cells = list(sheet.iter_rows(min_row=1, max_row=1))[0]

        # get columns
        columns = kwargs.pop("columns", None)
        columns_row = kwargs.pop("columns_row", True)
        columns_standardized = kwargs.pop("columns_standardized", columns is None)
        if not columns:
            if columns_row:
                # if first row is for column names read the names
                # for row in sheet.iter_rows(min_row=1, max_row=1):
                columns = [cell.value for cell in sheet_columns_cells]
            else:
                # otherwise use columns indexes as column names
                # for row in sheet.iter_rows(min_row=1, max_row=1):
                columns = self._get_sheet_columns_indexes(len(sheet_columns_cells))

        # standardize column names, eg. "Date Created" -> "date_created"
        if columns_standardized:
            columns = [slugify(column or "", separator="_") for column in columns]

        # build list of dicts, one for each row
        items = []
        items_row_start = 2 if columns_row else 1
        for row in sheet.iter_rows(min_row=items_row_start):
            values = [cell.value for cell in row]
            items.append(dict(zip(columns, values)))

        # close the worksheet
        workbook.close()

        # print(items)
        return items

    def decode(self, s, **kwargs):
        require_xls(installed=xls_installed)
        extension = fsutil.get_file_extension(s)
        if extension in ["xlsx", "xlsm"]:
            return self._decode(s, **kwargs)
        elif extension in ["xls", "xlt"]:
            return self._decode_legacy(s, **kwargs)

    def encode(self, d, **kwargs):
        # require_xls(installed=xls_installed)
        raise NotImplementedError
